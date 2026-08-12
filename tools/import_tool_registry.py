#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    sheet = workbook["Tools"]
    headers = [cell.value for cell in sheet[4]]
    rows = [
        {header: (value if value is not None else "") for header, value in zip(headers, values, strict=True)}
        for values in sheet.iter_rows(min_row=5, values_only=True)
        if any(value is not None for value in values)
    ]
    payload = {
        "schema_version": "0.1.0",
        "source_registry": "BRIDGE-P0-toolkit-master-20260810",
        "source_workbook_sha256": _sha256(args.workbook),
        "rows": rows,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return 0


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


if __name__ == "__main__":
    raise SystemExit(main())
